from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font

from app.load.func import *
from app.common.func import get_departments
from app.common.classes.EducationStaff import EducationStaff
from app.common.classes.ExcelStyle import ExcelStyle
from config import FlaskConfig, ApeksConfig as Apeks


class LoadData:
    def __init__(
        self,
        year: int,
        month: int,
        lesson_staff: dict,
        load_subgroup: dict,
        load_group: dict,
        plan_education_plans_education_form: dict,
        plan_education_plan: dict,
    ):
        self.year = year
        self.month = month
        self.departments = get_departments()
        self.lesson_staff = lesson_staff
        self.load_subgroups = load_subgroup
        self.load_groups = load_group
        self.plan_education_plans_education_forms = plan_education_plans_education_form
        self.plan_education_plans = plan_education_plan
        self.education_staff = EducationStaff(month, year)
        self.staff_dept_structure = self.staff_dept_structure()
        self.structured_lessons = self.process_lessons()
        self.control_lessons = self.control_lessons()

    def staff_dept_structure(self):
        """
        Department ID for prepod in current month
        {staff_id:department_id}.
        """
        structure = {}
        for dept in self.departments:
            staff_list = self.education_staff.department_staff(dept)
            for staff in staff_list:
                structure[staff] = dept
        return structure

    def process_lessons(self):
        """Current month lessons list with all data for load report"""
        structured_lessons = []
        if self.month == "январь-август":
            month_list = [1, 2, 3, 4, 5, 6, 7, 8]
        elif self.month == "сентябрь-декабрь":
            month_list = [9, 10, 11, 12]
        else:
            month_list = [int(self.month)]
        for month in month_list:
            for lesson in get_lessons(self.year, month):
                if not lesson.get("group_id"):
                    if lesson.get("subgroup_id"):
                        lesson["group_id"] = self.load_subgroups[
                            lesson.get("subgroup_id")
                        ].get("group_id")
                education_plan_id = self.load_groups.get(lesson.get("group_id"))[
                    "education_plan_id"
                ]
                education_form_id = self.plan_education_plans_education_forms.get(
                    education_plan_id
                )
                education_level_id = self.plan_education_plans.get(education_plan_id)[
                    "education_level_id"
                ]
                lesson["education_plan_id"] = education_plan_id
                lesson["education_form_id"] = education_form_id
                lesson["education_level_id"] = education_level_id
                if lesson.get("id") in self.lesson_staff:
                    for staff in self.lesson_staff[lesson["id"]]:
                        less_copy = copy(lesson)
                        less_copy["staff_id"] = int(staff)
                        less_copy["department_id"] = self.staff_dept_structure.get(
                            int(staff)
                        )
                        less_copy["hours"] = 2
                        structured_lessons.append(less_copy)
        return structured_lessons

    def control_lessons(self):
        """Получения занятий типа контроль (зачет, экзамен)"""

        def check_and_process(lesson):
            """Сверка типа занятия для определения типа контроль"""
            lookup = [
                "staff_id",
                "discipline_id",
                "group_id",
                "control_type_id",
                "date",
            ]
            counter = 0
            for c in control_less:
                if [lesson[val] for val in lookup] == [c[val] for val in lookup]:
                    c["hours"] += 2
                    c["lesson_time_id"] = (
                        c.get("lesson_time_id") + ", " + lesson.get("lesson_time_id")
                    )
                    counter += 1
            if counter != 0:
                return True
            else:
                return False

        control_less = []
        for less in self.structured_lessons:
            if less.get("control_type_id") in [
                str(Apeks.CONTROL_TYPE_ID.get(i))
                for i in ("exam", "zachet", "zachet_mark", "final_att", "kandidat_exam")
            ]:
                if not check_and_process(less):
                    control_less.append(copy(less))
        return control_less

    def get_control_hours(self, contr_less):
        """
        Getting hours for control lessons depending on students number,
        education type etc...
        Если значение больше максимального, возвращаем максимальное
        """
        cont_type = get_lesson_type(contr_less)
        stud_type = get_student_type(contr_less)
        if contr_less.get("subgroup_id"):  # If subgroup - get people count from it
            subgroup_id = contr_less.get("subgroup_id")
            people_count = self.load_subgroups[subgroup_id].get("people_count")
        else:
            group_id = contr_less.get("group_id")
            people_count = self.load_groups[group_id].get("people_count")
        if stud_type == "prof_p" or stud_type == "dpo":
            return contr_less["hours"]
        elif stud_type == "adj" and (
            contr_less.get("control_type_id") == Apeks.CONTROL_TYPE_ID.get("final_att")
            or contr_less.get("control_type_id")
            == Apeks.CONTROL_TYPE_ID.get("kandidat_exam")
        ):
            value = int(people_count) * Apeks.ADJ_KF
            return Apeks.ADJ_KF_MAX if value > Apeks.ADJ_KF_MAX else value
        elif cont_type == "zachet":
            value = int(people_count) * Apeks.ZACH_KF
            return Apeks.ZACH_KF_MAX if value > Apeks.ZACH_KF_MAX else value
        elif cont_type == "exam":
            value = int(people_count) * Apeks.EXAM_KF
            return Apeks.EXAM_KF_MAX if value > Apeks.EXAM_KF_MAX else value
        elif cont_type == "final_att":
            value = int(people_count) * Apeks.FINAL_KF
            return Apeks.FINAL_KF_MAX if value > Apeks.FINAL_KF_MAX else value

    def unknown_lessons(self):
        """
        Lessons with inactive staff
        (which doesn't work in department in selected time).
        """
        unknown = []
        for less in self.structured_lessons:
            if not less.get("department_id"):
                unknown.append(less)
        return unknown

    def department_lessons(self, department_id):
        """Select department lessons for load report"""
        dept_lessons = []
        for less in self.structured_lessons:
            if less.get("department_id") == int(department_id):
                dept_lessons.append(less)
        return dept_lessons


class DeptPrepodLoad:
    def __init__(self, staff_list: list):
        self.load = {}
        self.staff_list = staff_list
        lesson_types = [
            "lecture",
            "seminar",
            "pract",
            "group_cons",
            "zachet",
            "exam",
            "final_att",
        ]
        student_types = ["och", "zo_high", "zo_mid", "adj", "prof_p", "dpo"]
        for staff_id in staff_list:
            self.load[staff_id] = {}
            for l_type in lesson_types:
                self.load[staff_id][l_type] = {}
                for s_type in student_types:
                    self.load[staff_id][l_type][s_type] = 0

    def add_load(self, staff_id, l_type, s_type, value=2):
        self.load[staff_id][l_type][s_type] += value

    def get_load(self, staff_id):
        staff_load = {self.staff_list[staff_id]: self.load[staff_id]}
        return staff_load


class LoadReport:
    def __init__(self, year, month, department_id, load):
        self.year = year
        self.month = month
        self.department_id = int(department_id)
        self.staff_list = EducationStaff(month, year).department_staff(department_id)
        self.load = load
        self.dept_load = DeptPrepodLoad(list(self.staff_list))
        self.unprocessed = []
        for lesson in self.load.structured_lessons:
            if lesson.get("department_id") == self.department_id:
                staff_id = lesson.get("staff_id")
                l_type = get_lesson_type(lesson)
                s_type = get_student_type(lesson)
                if staff_id and l_type and s_type:
                    if l_type in ["lecture", "seminar", "pract", "group_cons"]:
                        self.dept_load.add_load(staff_id, l_type, s_type)
                else:
                    self.unprocessed.append(lesson)
        for control in self.load.control_lessons:
            if control.get("department_id") == self.department_id:
                staff_id = control.get("staff_id")
                l_type = get_lesson_type(control)
                s_type = get_student_type(control)
                if staff_id and l_type and s_type:
                    if l_type in ["zachet", "exam", "final_att"]:
                        value = self.load.get_control_hours(control)
                        self.dept_load.add_load(staff_id, l_type, s_type, value)
                else:
                    self.unprocessed.append(control)
        self.data = self.dept_load.load
        self.filename = (
            f"{self.year}-{self.month} "
            f"{self.load.departments.get(self.department_id).get('short')}.xlsx"
        )

    def generate_report(self):
        """Формирование отчета о нагрузке в Excel."""

        wb = load_workbook(FlaskConfig.TEMP_FILE_DIR + "load_report_temp.xlsx")
        ws = wb.active
        ws.title = (
            f"{self.year}-{self.month} "
            f"{self.load.departments.get(self.department_id).get('short')}"
        )
        ws.cell(1, 1).value = "кафедра " + self.load.departments.get(
            self.department_id
        ).get("full")
        ws.cell(2, 1).value = f"отчет о нагрузке за {self.month} - {self.year}"
        row = 8
        for prepod in self.data:
            # Style apply
            for i in range(2, 73):
                ws.cell(row, i).style = ExcelStyle.Number
            # Prepod Name
            ws.cell(row, 1).value = self.staff_list[prepod]
            ws.cell(row, 1).style = ExcelStyle.Base
            for l_type in self.data[prepod]:
                if l_type == "lecture":
                    column = 2
                elif l_type == "seminar":
                    column = 8
                elif l_type == "pract":
                    column = 14
                elif l_type == "group_cons":
                    column = 24
                    if self.data[prepod][l_type]["dpo"]:
                        del self.data[prepod][l_type]["dpo"]
                elif l_type == "zachet":
                    column = 29
                elif l_type == "exam":
                    column = 35
                elif l_type == "final_att":
                    column = 59
                else:
                    column = 73
                for key, val in self.data[prepod][l_type].items():
                    val = "" if val == 0 else val
                    ws.cell(row, column).value = val
                    if val and val % 1 > 0:
                        ws.cell(row, column).number_format = "0.00"
                    column += 1
            ws.cell(row, 72).value = f"=SUM(B{str(row)}:BS{str(row)})"
            row += 1
        # Total
        ws.cell(row, 1).value = "Итого"
        ws.cell(row, 1).style = ExcelStyle.BaseBold
        for col in range(2, 73):
            ltr = ws.cell(row, col).column_letter
            ws.cell(row, col).value = (
                f"=IF(SUM({ltr}8:{ltr}{str(row - 1)})>0,"
                f'SUM({ltr}8:{ltr}{str(row - 1)}),"")'
            )
            ws.cell(row, col).style = ExcelStyle.Number
            ws.cell(row, col).font = Font(name="Times New Roman", size=10, bold=True)
        wb.save(FlaskConfig.EXPORT_FILE_DIR + self.filename)
