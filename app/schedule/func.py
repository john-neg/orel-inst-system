from __future__ import annotations

import logging
from datetime import datetime, timedelta

import pytz
from icalendar import Calendar, Timezone, TimezoneStandard, Event, Alarm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.common.classes.ExcelStyle import ExcelStyle
from app.common.classes.ScheduleLessonsStaff import ScheduleLessonsStaff
from config import FlaskConfig, ApeksConfig as Apeks


def lessons_ical_exp(
    staff_id: int | str,
    staff_name: str,
    month: int | str,
    year: int | str,
    timezone=Apeks.TIMEZONE,
) -> str:
    """Формирование файла для экспорта занятий преподавателя в формате iCal."""
    lessons = ScheduleLessonsStaff(staff_id, month, year)

    if not lessons.data:
        return "no data"

    cal = Calendar()
    cal.add("calscale", "GREGORIAN")
    cal.add("version", "2.0")
    cal.add("prodid", "APEKS-VUZ-EXTENSION")
    cal_timezone = Timezone()
    cal_timezone.add("TZID", Apeks.TIMEZONE)
    tz_standard = TimezoneStandard()
    tz_standard.add("TZNAME", datetime.now(tz=timezone).strftime("%Z"))
    tz_standard["TZOFFSETFROM"] = datetime.now(tz=timezone).strftime("%z")
    tz_standard["TZOFFSETTO"] = datetime.now(tz=timezone).strftime("%z")
    cal_timezone.add_component(tz_standard)
    cal.add_component(cal_timezone)

    for l_index in range(len(lessons.data)):
        event = Event()
        event.add("dtstart", lessons.time_start(l_index).astimezone(pytz.utc))
        event.add("dtend", lessons.time_end(l_index).astimezone(pytz.utc))
        event.add("dtstamp", datetime.now().astimezone(pytz.utc))
        event.add(
            "description",
            f"т.{lessons.data[l_index].get('topic_code')} "
            f"{lessons.data[l_index].get('topic_name')}\n\n"
            f"Данные актуальны на: {datetime.now().strftime('%H:%M %d.%m.%Y')}",
        )
        event.add("location", lessons.data[l_index].get("classroom"))
        event.add("status", "CONFIRMED")
        event.add("summary", lessons.calendar_name(l_index))
        event.add(
            "uid",
            f'apeks-id-{lessons.data[l_index].get("id")}'
            f'journal-lesson-id-{lessons.data[l_index].get("journal_lesson_id")}',
        )
        alarm = Alarm()
        alarm.add("action", "DISPLAY")
        alarm.add("description", "Напоминание")
        alarm.add("trigger", timedelta(minutes=-30))
        event.add_component(alarm)
        cal.add_component(event)

    filename = f"{staff_name} {month}-{year}.ics"

    with open(
        f"{FlaskConfig.EXPORT_FILE_DIR}{filename}",
        "wb",
    ) as f:
        f.write(cal.to_ical())
        f.close()
        logging.debug(f'Файл "{filename}" успешно сформирован')
    return filename


def lessons_xlsx_exp(
    staff_id: int | str, staff_name: str, month: int | str, year: int | str
) -> str:
    """Формирование файла для экспорта занятий преподавателя в формате xlsx."""
    lessons = ScheduleLessonsStaff(staff_id, month, year)

    if not lessons.data:
        return "no data"
    else:
        filename = f"{staff_name} {month}-{year}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = staff_name

        row = 1
        ws.cell(row, 1).value = staff_name
        ws.cell(row, 1).style = ExcelStyle.Header
        ws.cell(row, 2).value = f"Расписание на месяц {str(month)}-{str(year)}"
        ws.cell(row, 2).style = ExcelStyle.Header

        row = 2
        column = 1
        # Заголовки и ширина столбца
        headers = {"Дата/время": 15, "Занятие": 65, "Место": 15, "Тема": 80}
        for key, val in headers.items():
            ws.cell(row, column).value = key
            ws.cell(row, column).style = ExcelStyle.Header
            ws.cell(row, column).fill = ExcelStyle.GreyFill
            ws.column_dimensions[get_column_letter(column)].width = val
            column += 1

        row = 3
        for l_index in range(len(lessons.data)):
            ws.cell(row, 1).value = lessons.time_start(l_index).strftime(
                "%d.%m.%Y %H:%M"
            )
            ws.cell(row, 1).style = ExcelStyle.Base_No_Wrap
            ws.cell(row, 2).value = lessons.calendar_name(l_index)
            ws.cell(row, 2).style = ExcelStyle.Base_No_Wrap
            ws.cell(row, 3).value = lessons.data[l_index].get("classroom")
            ws.cell(row, 3).style = ExcelStyle.Base_No_Wrap
            ws.cell(row, 4).value = lessons.data[l_index].get("topic_name")
            ws.cell(row, 4).style = ExcelStyle.Base_No_Wrap
            row += 1

        wb.save(FlaskConfig.EXPORT_FILE_DIR + filename)
        logging.debug(f'Файл "{filename}" успешно сформирован')
        return filename
