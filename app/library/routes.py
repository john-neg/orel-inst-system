import os

from flask import render_template, request, redirect, url_for, flash
from flask.views import View
from flask_login import login_required
from openpyxl import load_workbook

from app.library import bp
from app.library.func import library_file_processing, load_bibl
from app.library.models import LibraryPlan
from app.main.forms import ChoosePlan, FileForm
from app.main.func import (
    education_specialty,
    education_plans,
    db_filter_req,
    allowed_file,
)
from config import FlaskConfig, LibConfig


LIB_TYPES = {
    "library": [LibConfig.BIBL_MAIN, LibConfig.BIBL_ADD],
    "library_np": [LibConfig.BIBL_NP],
    "library_int": [LibConfig.BIBL_INT],
    "library_db": [LibConfig.BIBL_DB],
}


class ChoosePlanView(View):
    methods = ["GET", "POST"]

    def __init__(self, lib_type, title, lib_type_name):
        self.template_name = "library/library_choose_plan.html"
        self.lib_type = lib_type
        self.lib_type_name = lib_type_name
        self.title = title

    @login_required
    def dispatch_request(self):
        form = ChoosePlan()
        form.edu_spec.choices = list(education_specialty().items())
        if request.method == "POST":
            edu_spec = request.form.get("edu_spec")
            form.edu_plan.choices = list(education_plans(edu_spec).items())
            if request.form.get("edu_plan") and form.validate_on_submit():
                edu_plan = request.form.get("edu_plan")
                return redirect(
                    url_for(
                        f"library.{self.lib_type}_upload",
                        plan_id=edu_plan,
                    )
                )
            return render_template(
                self.template_name,
                active="library",
                title=self.title,
                form=form,
                edu_spec=edu_spec,
            )
        return render_template(
            self.template_name,
            active="library",
            title=self.title,
            form=form,
        )


bp.add_url_rule(
    "/library_choose_plan",
    view_func=ChoosePlanView.as_view(
        "library_choose_plan",
        lib_type="library",
        lib_type_name="Литература",
        title="Загрузка списка литературы",
    ),
)
bp.add_url_rule(
    "/library_np_choose_plan",
    view_func=ChoosePlanView.as_view(
        "library_np_choose_plan",
        lib_type="library_np",
        lib_type_name="Научная продукция",
        title="Загрузка списка научной продукции",
    ),
)
bp.add_url_rule(
    "/library_int_choose_plan",
    view_func=ChoosePlanView.as_view(
        "library_int_choose_plan",
        lib_type="library_int",
        lib_type_name="Интернет ресурсы",
        title="Загрузка ресурсов сети Интернет",
    ),
)
bp.add_url_rule(
    "/library_db_choose_plan",
    view_func=ChoosePlanView.as_view(
        "library_db_choose_plan",
        lib_type="library_db",
        lib_type_name="Базы и справочные системы",
        title="Загрузка ресурсов баз данных и инф.-справ. систем",
    ),
)


class LibraryUploadView(View):
    methods = ["GET", "POST"]

    def __init__(self, lib_type, lib_type_name, title):
        self.template_name = "library/library_upload.html"
        self.lib_type = lib_type
        self.lib_type_name = lib_type_name
        self.title = title

    @login_required
    def dispatch_request(self, plan_id):
        form = FileForm()
        plan = LibraryPlan(plan_id)
        if request.method == "POST":
            if request.form.get("library_load_temp"):
                return redirect(
                    url_for(
                        "main.get_temp_file", filename=f"{self.lib_type}_load_temp.xlsx"
                    )
                )
            if request.form.get("library_plan_content"):
                return redirect(
                    url_for(f"library.{self.lib_type}_export", plan_id=plan_id)
                )
            if request.files["file"]:
                file = request.files["file"]
                if file and allowed_file(file.filename):
                    filename = file.filename
                    file.save(os.path.join(FlaskConfig.UPLOAD_FILE_DIR, filename))
                    if request.form.get("library_check"):
                        # Проверка данных
                        return redirect(
                            url_for(
                                f"library.{self.lib_type}_check",
                                plan_id=plan_id,
                                filename=filename,
                            )
                        )
                    if request.form.get("library_update"):
                        # Загрузка данных
                        return redirect(
                            url_for(
                                f"library.{self.lib_type}_update",
                                plan_id=plan_id,
                                filename=filename,
                            )
                        )
        return render_template(
            self.template_name,
            active="library",
            title=self.title,
            lib_type_name=self.lib_type_name,
            form=form,
            plan_name=plan.name,
        )


bp.add_url_rule(
    "/library_upload/<int:plan_id>",
    view_func=LibraryUploadView.as_view(
        "library_upload",
        lib_type="library",
        lib_type_name="Литература",
        title="Загрузка списка литературы",
    ),
)
bp.add_url_rule(
    "/library_np_upload/<int:plan_id>",
    view_func=LibraryUploadView.as_view(
        "library_np_upload",
        lib_type="library_np",
        lib_type_name="Научная продукция",
        title="Загрузка списка научной продукции",
    ),
)
bp.add_url_rule(
    "/library_int_upload/<int:plan_id>",
    view_func=LibraryUploadView.as_view(
        "library_int_upload",
        lib_type="library_int",
        lib_type_name="Интернет ресурсы",
        title="Загрузка ресурсов сети Интернет",
    ),
)
bp.add_url_rule(
    "/library_db_upload/<int:plan_id>",
    view_func=LibraryUploadView.as_view(
        "library_db_upload",
        lib_type="library_db",
        lib_type_name="Базы и справочные системы",
        title="Загрузка ресурсов баз данных и инф.-справ. систем",
    ),
)


class LibraryCheckView(View):
    methods = ["GET", "POST"]

    def __init__(self, lib_type, lib_type_name, title):
        self.template_name = "library/library_upload.html"
        self.lib_type = lib_type
        self.lib_type_name = lib_type_name
        self.title = title

    @login_required
    def dispatch_request(self, plan_id, filename):
        file = FlaskConfig.UPLOAD_FILE_DIR + filename
        form = FileForm()
        plan = LibraryPlan(plan_id)
        lib_data = library_file_processing(file)
        if request.method == "POST":
            if request.files["file"]:
                file = request.files["file"]
                if file and allowed_file(file.filename):
                    filename = file.filename
                    file.save(os.path.join(FlaskConfig.UPLOAD_FILE_DIR, filename))
                    if request.form.get("library_check"):
                        return redirect(
                            url_for(
                                f"library.{self.lib_type}_check",
                                plan_id=plan_id,
                                filename=filename,
                            )
                        )
            if request.form.get("library_load_temp"):
                return redirect(
                    url_for(
                        "main.get_temp_file", filename=f"{self.lib_type}_load_temp.xlsx"
                    )
                )
            if request.form.get("library_plan_content"):
                return redirect(
                    url_for(f"library.{self.lib_type}_export", plan_id=plan_id)
                )
            if request.form.get("library_update"):
                return redirect(
                    url_for(
                        f"library.{self.lib_type}_update",
                        plan_id=plan_id,
                        filename=filename,
                    )
                )
        # Check if program in uploaded file
        work_programs, no_data = [], []
        for wp_id in plan.work_programs:
            if plan.work_programs.get(wp_id) in lib_data:
                work_programs.append(plan.work_programs.get(wp_id))
            else:
                no_data.append(plan.work_programs.get(wp_id))
        return render_template(
            self.template_name,
            active="library",
            title=self.title,
            lib_type_name=self.lib_type_name,
            form=form,
            plan_name=plan.name,
            no_data=no_data,
            no_program=plan.non_exist,
            work_programs=work_programs,
        )


bp.add_url_rule(
    "/library_check/<int:plan_id>/<string:filename>",
    view_func=LibraryCheckView.as_view(
        "library_check",
        lib_type="library",
        lib_type_name="Литература",
        title="Загрузка списка литературы",
    ),
)
bp.add_url_rule(
    "/library_np_check/<int:plan_id>/<string:filename>",
    view_func=LibraryCheckView.as_view(
        "library_np_check",
        lib_type="library_np",
        lib_type_name="Научная продукция",
        title="Загрузка списка научной продукции",
    ),
)
bp.add_url_rule(
    "/library_int_check/<int:plan_id>/<string:filename>",
    view_func=LibraryCheckView.as_view(
        "library_int_check",
        lib_type="library_int",
        lib_type_name="Интернет ресурсы",
        title="Загрузка ресурсов сети Интернет",
    ),
)
bp.add_url_rule(
    "/library_db_check/<int:plan_id>/<string:filename>",
    view_func=LibraryCheckView.as_view(
        "library_db_check",
        lib_type="library_db",
        lib_type_name="Базы и справочные системы",
        title="Загрузка ресурсов баз данных и инф.-справ. систем",
    ),
)


class LibraryUpdateView(View):
    methods = ["GET", "POST"]

    def __init__(self, lib_type, lib_type_name):
        self.lib_type = lib_type
        self.lib_type_name = lib_type_name

    @login_required
    def dispatch_request(self, plan_id, filename):
        file = FlaskConfig.UPLOAD_FILE_DIR + filename
        plan = LibraryPlan(plan_id)
        file_data = library_file_processing(file)
        for disc in file_data:
            for wp_id in plan.work_programs:
                if plan.work_programs.get(wp_id) == disc:
                    counter = 0
                    for bibl in LIB_TYPES[self.lib_type]:
                        load_bibl(wp_id, bibl, file_data[disc][counter])
                    counter += 1
        flash(f"Данные из файла - {filename}: успешно загружены")
        return redirect(url_for(f"library.{self.lib_type}_upload", plan_id=plan_id))


bp.add_url_rule(
    "/library_update/<int:plan_id>/<string:filename>",
    view_func=LibraryUpdateView.as_view(
        "library_update",
        lib_type="library",
        lib_type_name="Литература",
    ),
)
bp.add_url_rule(
    "/library_np_update/<int:plan_id>/<string:filename>",
    view_func=LibraryUpdateView.as_view(
        "library_np_update",
        lib_type="library_np",
        lib_type_name="Научная продукция",
    ),
)
bp.add_url_rule(
    "/library_int_update/<int:plan_id>/<string:filename>",
    view_func=LibraryUpdateView.as_view(
        "library_int_update",
        lib_type="library_int",
        lib_type_name="Интернет ресурсы",
    ),
)
bp.add_url_rule(
    "/library_db_update/<int:plan_id>/<string:filename>",
    view_func=LibraryUpdateView.as_view(
        "library_db_update",
        lib_type="library_db",
        lib_type_name="Базы и справочные системы",
    ),
)


class LibraryExportView(View):
    methods = ["GET", "POST"]

    def __init__(self, lib_type, lib_type_name):
        self.lib_type = lib_type
        self.lib_type_name = lib_type_name

    @login_required
    def dispatch_request(self, plan_id):
        plan = LibraryPlan(plan_id)
        lib_data = plan.library_content()
        filename = f'{self.lib_type_name} - {db_filter_req("plan_education_plans", "id", plan_id)[0]["name"]}.xlsx'
        wb = load_workbook(
            FlaskConfig.TEMP_FILE_DIR + f"{self.lib_type}_load_temp.xlsx"
        )
        ws = wb.active
        start_row = 2
        for data in lib_data:
            ws.cell(row=start_row, column=1).value = data
            counter = 0
            for bibl in LIB_TYPES[self.lib_type]:
                ws.cell(row=start_row, column=counter + 2).value = lib_data[data][bibl]
                counter += 1
            start_row += 1
        wb.save(FlaskConfig.EXPORT_FILE_DIR + filename)
        return redirect(url_for("main.get_file", filename=filename))


bp.add_url_rule(
    "/library_export/<int:plan_id>",
    view_func=LibraryExportView.as_view(
        "library_export",
        lib_type="library",
        lib_type_name="Литература",
    ),
)
bp.add_url_rule(
    "/library_np_export/<int:plan_id>",
    view_func=LibraryExportView.as_view(
        "library_np_export",
        lib_type="library_np",
        lib_type_name="Научная продукция",
    ),
)
bp.add_url_rule(
    "/library_int_export/<int:plan_id>",
    view_func=LibraryExportView.as_view(
        "library_int_export",
        lib_type="library_int",
        lib_type_name="Интернет ресурсы",
    ),
)
bp.add_url_rule(
    "/library_db_export/<int:plan_id>",
    view_func=LibraryExportView.as_view(
        "library_db_export",
        lib_type="library_db",
        lib_type_name="Базы и справочные системы",
    ),
)
