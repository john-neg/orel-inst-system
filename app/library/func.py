from __future__ import annotations

from openpyxl import load_workbook

from app.common.func.app_core import xlsx_iter_rows, xlsx_normalize


def library_file_processing(file: str) -> dict:
    """
    Обработка загруженного файла c данными об обеспечении.
    (словарь без первой строки файла).

    Parameters
    ----------
        file: str
            полный путь к файлу с обеспечением для дисциплины

    Returns
    -------
        dict
            {"Название дисциплины": ["данные об обеспечении"]}

    """
    wb = load_workbook(file)
    ws = wb.active
    replace_dict = {
        "  ": " ",
        "–": "-",
        "\t": "",
        "_x000D_": "",
        "None": "",
        "Нет программы": ""
    }
    ws = xlsx_normalize(ws, replace_dict)
    lib_list = list(xlsx_iter_rows(ws))
    del lib_list[0]
    lib_dict = {}
    for lib in lib_list:
        lib_dict[lib[0].strip()] = []
        for i in range(1, len(lib)):
            lib_dict[lib[0]].append(lib[i])
    return lib_dict
