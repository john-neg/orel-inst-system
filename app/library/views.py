import os

from flask import render_template, request, redirect, url_for, flash
from flask.views import View
from flask_login import login_required
from openpyxl import load_workbook

from app.common.classes.EducationPlan import EducationPlan, EducationPlanWorkProgram
from app.common.forms import ChoosePlan, FileForm
from app.common.func import (
    check_api_db_response,
    api_get_db_table,
    get_plan_curriculum_disciplines,
    get_work_programs_data,
    allowed_file,
    get_plan_education_specialties,
    get_education_plans,
)
from app.library import bp
from app.library.func import library_file_processing, load_bibl
from app.library.models import LibraryPlan
from config import FlaskConfig, ApeksConfig as Apeks

LIB_TYPES = {
    "library": [
        Apeks.MM_WORK_PROGRAMS_DATA_ITEMS.get("library_main"),
        Apeks.MM_WORK_PROGRAMS_DATA_ITEMS.get("library_add"),
    ],
    "library_np": [Apeks.MM_WORK_PROGRAMS_DATA_ITEMS.get("library_np")],
    "library_int": [Apeks.MM_WORK_PROGRAMS_DATA_ITEMS.get("internet")],
    "library_db": [Apeks.MM_WORK_PROGRAMS_DATA_ITEMS.get("ref_system")],
}


class LibraryChoosePlanView(View):
    methods = ["GET", "POST"]

    def __init__(self, lib_type, title, lib_type_name):
        self.template_name = "library/library_choose_plan.html"
        self.lib_type = lib_type
        self.lib_type_name = lib_type_name
        self.title = title

    @login_required
    async def dispatch_request(self):
        form = ChoosePlan()
        specialities = await get_plan_education_specialties()
        form.edu_spec.choices = list(specialities.items())
        if request.method == "POST":
            edu_spec = request.form.get("edu_spec")
            plans = await get_education_plans(edu_spec)
            form.edu_plan.choices = list(plans.items())
            if request.form.get("edu_plan") and form.validate_on_submit():
                edu_plan = request.form.get("edu_plan")
                return redirect(
                    url_for(
                        f"library.{self.lib_type}_upload",
                        plan_id=edu_plan,
                    )
                )
            return render_template(
                self.template_name,
                active="library",
                title=self.title,
                form=form,
                edu_spec=edu_spec,
            )
        return render_template(
            self.template_name,
            active="library",
            title=self.title,
            form=form,
        )


bp.add_url_rule(
    "/library_choose_plan",
    view_func=LibraryChoosePlanView.as_view(
        "library_choose_plan",
        lib_type="library",
        lib_type_name="Литература",
        title="Загрузка списка литературы",
    ),
)
bp.add_url_rule(
    "/library_np_choose_plan",
    view_func=LibraryChoosePlanView.as_view(
        "library_np_choose_plan",
        lib_type="library_np",
        lib_type_name="Научная продукция",
        title="Загрузка списка научной продукции",
    ),
)
bp.add_url_rule(
    "/library_int_choose_plan",
    view_func=LibraryChoosePlanView.as_view(
        "library_int_choose_plan",
        lib_type="library_int",
        lib_type_name="Интернет ресурсы",
        title="Загрузка ресурсов сети Интернет",
    ),
)
bp.add_url_rule(
    "/library_db_choose_plan",
    view_func=LibraryChoosePlanView.as_view(
        "library_db_choose_plan",
        lib_type="library_db",
        lib_type_name="Базы и справочные системы",
        title="Загрузка ресурсов баз данных и инф.-справ. систем",
    ),
)


class LibraryUploadView(View):
    methods = ["GET", "POST"]

    def __init__(self, lib_type, lib_type_name, title):
        self.template_name = "library/library_upload.html"
        self.lib_type = lib_type
        self.lib_type_name = lib_type_name
        self.title = title

    @login_required
    async def dispatch_request(self, plan_id):
        form = FileForm()
        plan = EducationPlan(
            education_plan_id=plan_id,
            plan_education_plans=await check_api_db_response(
                await api_get_db_table(
                    Apeks.TABLES.get("plan_education_plans"), id=plan_id
                )
            ),
            plan_curriculum_disciplines=await get_plan_curriculum_disciplines(plan_id),
        )
        if request.method == "POST":
            if request.form.get("library_load_temp"):
                return redirect(
                    url_for(
                        "main.get_temp_file", filename=f"{self.lib_type}_load_temp.xlsx"
                    )
                )
            if request.form.get("library_plan_content"):
                return redirect(
                    url_for(f"library.{self.lib_type}_export", plan_id=plan_id)
                )
            if request.files["file"]:
                file = request.files["file"]
                if file and allowed_file(file.filename):
                    filename = file.filename
                    file.save(os.path.join(FlaskConfig.UPLOAD_FILE_DIR, filename))
                    if request.form.get("library_check"):
                        # Проверка данных
                        return redirect(
                            url_for(
                                f"library.{self.lib_type}_check",
                                plan_id=plan_id,
                                filename=filename,
                            )
                        )
                    if request.form.get("library_update"):
                        # Загрузка данных
                        return redirect(
                            url_for(
                                f"library.{self.lib_type}_update",
                                plan_id=plan_id,
                                filename=filename,
                            )
                        )
        return render_template(
            self.template_name,
            active="library",
            title=self.title,
            lib_type_name=self.lib_type_name,
            form=form,
            plan_name=plan.name,
        )


bp.add_url_rule(
    "/library_upload/<int:plan_id>",
    view_func=LibraryUploadView.as_view(
        "library_upload",
        lib_type="library",
        lib_type_name="Литература",
        title="Загрузка списка литературы",
    ),
)
bp.add_url_rule(
    "/library_np_upload/<int:plan_id>",
    view_func=LibraryUploadView.as_view(
        "library_np_upload",
        lib_type="library_np",
        lib_type_name="Научная продукция",
        title="Загрузка списка научной продукции",
    ),
)
bp.add_url_rule(
    "/library_int_upload/<int:plan_id>",
    view_func=LibraryUploadView.as_view(
        "library_int_upload",
        lib_type="library_int",
        lib_type_name="Интернет ресурсы",
        title="Загрузка ресурсов сети Интернет",
    ),
)
bp.add_url_rule(
    "/library_db_upload/<int:plan_id>",
    view_func=LibraryUploadView.as_view(
        "library_db_upload",
        lib_type="library_db",
        lib_type_name="Базы и справочные системы",
        title="Загрузка ресурсов баз данных и инф.-справ. систем",
    ),
)


class LibraryCheckView(View):
    methods = ["GET", "POST"]

    def __init__(self, lib_type, lib_type_name, title):
        self.template_name = "library/library_upload.html"
        self.lib_type = lib_type
        self.lib_type_name = lib_type_name
        self.title = title

    @login_required
    async def dispatch_request(self, plan_id, filename):
        file = FlaskConfig.UPLOAD_FILE_DIR + filename
        form = FileForm()
        plan_disciplines = await get_plan_curriculum_disciplines(plan_id)
        plan = EducationPlanWorkProgram(
            education_plan_id=plan_id,
            plan_education_plans=await check_api_db_response(
                await api_get_db_table(
                    Apeks.TABLES.get("plan_education_plans"), id=plan_id
                )
            ),
            plan_curriculum_disciplines=plan_disciplines,
            work_programs_data=await get_work_programs_data([*plan_disciplines]),
        )
        lib_data = library_file_processing(file)
        if request.method == "POST":
            if request.files["file"]:
                file = request.files["file"]
                if file and allowed_file(file.filename):
                    filename = file.filename
                    file.save(os.path.join(FlaskConfig.UPLOAD_FILE_DIR, filename))
                    if request.form.get("library_check"):
                        return redirect(
                            url_for(
                                f"library.{self.lib_type}_check",
                                plan_id=plan_id,
                                filename=filename,
                            )
                        )
            if request.form.get("library_load_temp"):
                return redirect(
                    url_for(
                        "main.get_temp_file", filename=f"{self.lib_type}_load_temp.xlsx"
                    )
                )
            if request.form.get("library_plan_content"):
                return redirect(
                    url_for(f"library.{self.lib_type}_export", plan_id=plan_id)
                )
            if request.form.get("library_update"):
                return redirect(
                    url_for(
                        f"library.{self.lib_type}_update",
                        plan_id=plan_id,
                        filename=filename,
                    )
                )
        # Check if program in uploaded file
        work_programs, no_data = [], []
        for wp in plan.work_programs_data:
            wp_name = plan.work_programs_data[wp].get("name").strip()
            if wp_name in lib_data:
                work_programs.append(wp_name)
            else:
                no_data.append(wp_name)
        return render_template(
            self.template_name,
            active="library",
            title=self.title,
            lib_type_name=self.lib_type_name,
            form=form,
            plan_name=plan.name,
            no_data=no_data,
            no_program=plan.non_exist,
            duplicate=plan.duplicate,
            wrong_name=plan.wrong_name,
            work_programs=work_programs,
        )


bp.add_url_rule(
    "/library_check/<int:plan_id>/<string:filename>",
    view_func=LibraryCheckView.as_view(
        "library_check",
        lib_type="library",
        lib_type_name="Литература",
        title="Загрузка списка литературы",
    ),
)
bp.add_url_rule(
    "/library_np_check/<int:plan_id>/<string:filename>",
    view_func=LibraryCheckView.as_view(
        "library_np_check",
        lib_type="library_np",
        lib_type_name="Научная продукция",
        title="Загрузка списка научной продукции",
    ),
)
bp.add_url_rule(
    "/library_int_check/<int:plan_id>/<string:filename>",
    view_func=LibraryCheckView.as_view(
        "library_int_check",
        lib_type="library_int",
        lib_type_name="Интернет ресурсы",
        title="Загрузка ресурсов сети Интернет",
    ),
)
bp.add_url_rule(
    "/library_db_check/<int:plan_id>/<string:filename>",
    view_func=LibraryCheckView.as_view(
        "library_db_check",
        lib_type="library_db",
        lib_type_name="Базы и справочные системы",
        title="Загрузка ресурсов баз данных и инф.-справ. систем",
    ),
)


class LibraryUpdateView(View):
    methods = ["GET", "POST"]

    def __init__(self, lib_type, lib_type_name):
        self.lib_type = lib_type
        self.lib_type_name = lib_type_name

    @login_required
    def dispatch_request(self, plan_id, filename):
        file = FlaskConfig.UPLOAD_FILE_DIR + filename
        plan = LibraryPlan(plan_id)
        file_data = library_file_processing(file)
        for disc in file_data:
            for wp_id in plan.work_programs:
                if plan.work_programs.get(wp_id) == disc:
                    counter = 0
                    # TODO Добавить проверку существования поля и если нет то создавать
                    for bibl_type in LIB_TYPES[self.lib_type]:
                        load_bibl(wp_id, bibl_type, file_data[disc][counter])
                        counter += 1
        flash(f"Данные из файла - {filename}: успешно загружены")
        return redirect(url_for(f"library.{self.lib_type}_upload", plan_id=plan_id))


bp.add_url_rule(
    "/library_update/<int:plan_id>/<string:filename>",
    view_func=LibraryUpdateView.as_view(
        "library_update",
        lib_type="library",
        lib_type_name="Литература",
    ),
)
bp.add_url_rule(
    "/library_np_update/<int:plan_id>/<string:filename>",
    view_func=LibraryUpdateView.as_view(
        "library_np_update",
        lib_type="library_np",
        lib_type_name="Научная продукция",
    ),
)
bp.add_url_rule(
    "/library_int_update/<int:plan_id>/<string:filename>",
    view_func=LibraryUpdateView.as_view(
        "library_int_update",
        lib_type="library_int",
        lib_type_name="Интернет ресурсы",
    ),
)
bp.add_url_rule(
    "/library_db_update/<int:plan_id>/<string:filename>",
    view_func=LibraryUpdateView.as_view(
        "library_db_update",
        lib_type="library_db",
        lib_type_name="Базы и справочные системы",
    ),
)


class LibraryExportView(View):
    methods = ["GET", "POST"]

    def __init__(self, lib_type, lib_type_name):
        self.lib_type = lib_type
        self.lib_type_name = lib_type_name

    @login_required
    async def dispatch_request(self, plan_id):
        plan_disciplines = await get_plan_curriculum_disciplines(plan_id)
        plan = EducationPlanWorkProgram(
            education_plan_id=plan_id,
            plan_education_plans=await check_api_db_response(
                await api_get_db_table(
                    Apeks.TABLES.get("plan_education_plans"), id=plan_id
                )
            ),
            plan_curriculum_disciplines=plan_disciplines,
            work_programs_data=await get_work_programs_data(
                [*plan_disciplines], fields=True
            ),
        )
        lib_data = plan.library_content()
        filename = (
            f"{self.lib_type_name} - {plan.name}.xlsx"
        )
        wb = load_workbook(
            FlaskConfig.TEMPLATE_FILE_DIR + f"{self.lib_type}_load_temp.xlsx"
        )
        ws = wb.active
        start_row = 2
        for data in lib_data:
            ws.cell(row=start_row, column=1).value = data
            counter = 0
            for bibl in LIB_TYPES[self.lib_type]:
                ws.cell(row=start_row, column=counter + 2).value = lib_data[data][bibl]
                counter += 1
            start_row += 1
        wb.save(FlaskConfig.EXPORT_FILE_DIR + filename)
        return redirect(url_for("main.get_file", filename=filename))


bp.add_url_rule(
    "/library_export/<int:plan_id>",
    view_func=LibraryExportView.as_view(
        "library_export",
        lib_type="library",
        lib_type_name="Литература",
    ),
)
bp.add_url_rule(
    "/library_np_export/<int:plan_id>",
    view_func=LibraryExportView.as_view(
        "library_np_export",
        lib_type="library_np",
        lib_type_name="Научная продукция",
    ),
)
bp.add_url_rule(
    "/library_int_export/<int:plan_id>",
    view_func=LibraryExportView.as_view(
        "library_int_export",
        lib_type="library_int",
        lib_type_name="Интернет ресурсы",
    ),
)
bp.add_url_rule(
    "/library_db_export/<int:plan_id>",
    view_func=LibraryExportView.as_view(
        "library_db_export",
        lib_type="library_db",
        lib_type_name="Базы и справочные системы",
    ),
)
