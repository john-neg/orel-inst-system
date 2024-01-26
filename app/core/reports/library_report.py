import logging
import os

from openpyxl import load_workbook

from config import FlaskConfig, ApeksConfig as Apeks


def library_report(lib_data: dict, lib_type: str, filename: str) -> None:
    wb = load_workbook(
        os.path.join(FlaskConfig.TEMPLATE_FILE_DIR, f"{lib_type}_load_temp.xlsx")
    )
    ws = wb.active
    start_row = 2
    for data in lib_data:
        ws.cell(row=start_row, column=1).value = data
        counter = 0
        for bibl in Apeks.LIB_TYPES[lib_type]:
            ws.cell(row=start_row, column=counter + 2).value = lib_data[data][bibl]
            counter += 1
        start_row += 1
    wb.save(os.path.join(FlaskConfig.EXPORT_FILE_DIR, filename))
    logging.debug(f"Сформирован файл: {filename}")
