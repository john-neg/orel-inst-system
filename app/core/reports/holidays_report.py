import os
from datetime import date

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from config import ApeksConfig as Apeks, FlaskConfig
from ..classes.EducationStaff import EducationStaff
from ..func.api_get import api_get_db_table, check_api_db_response, get_lessons
from ..func.staff import get_state_staff
from ..reports.ExcelStyles import ExcelStyle
from ..services.apeks_db_state_departments_service import get_db_apeks_state_departments_service


def is_holiday(
    check_date: date, working_sat: list[date], non_working: list[date]
) -> bool:
    return (
        check_date.isoweekday() in [6, 7]
        and check_date not in working_sat
        or check_date in non_working
    )


async def generate_holidays_report(
    year: int,
    month_start: int,
    month_end: int,
    working_sat: list[date],
    non_working: list[date],
) -> str:
    """Формирует отчет о занятости в выходные в формате xlsx."""

    schedule_lessons_staff = await check_api_db_response(
        await api_get_db_table(
            Apeks.TABLES.get("schedule_day_schedule_lessons_staff"),
        )
    )

    lessons_staff = {}

    for lesson in schedule_lessons_staff:
        lesson_id = lesson.get("lesson_id")
        staff = lessons_staff.setdefault(lesson_id, [])
        staff.append(lesson.get("staff_id"))

    departments_service = get_db_apeks_state_departments_service()

    all_staff = EducationStaff(
        year,
        month_start,
        month_end,
        state_staff=await get_state_staff(),
        state_staff_history=await check_api_db_response(
            await api_get_db_table(Apeks.TABLES.get("state_staff_history"))
        ),
        state_staff_positions=await check_api_db_response(
            await api_get_db_table(Apeks.TABLES.get("state_staff_positions"))
        ),
        departments=await departments_service.get_departments(department_filter="kafedra"),
    )

    schedule_lessons = await check_api_db_response(
        await get_lessons(year, month_start, month_end)
    )

    staff_busy_holidays = {}
    total_holidays = set()

    for lesson in schedule_lessons:
        lesson_date = date.fromisoformat(lesson.get("date"))
        lesson_id = lesson.get("id")
        lesson_staff = lessons_staff.get(lesson_id)
        if lesson_staff:
            for staff_id in lesson_staff:
                staff_data = staff_busy_holidays.setdefault(
                    int(staff_id),
                    {
                        "name": all_staff.state_staff.get(int(staff_id)).get("full"),
                        "busy_holidays": set(),
                        "total_lessons": 0,
                        "holidays_lessons": 0,
                    },
                )
                staff_data["total_lessons"] += 1
                if is_holiday(lesson_date, working_sat, non_working):
                    total_holidays.add(lesson_date)
                    staff_data["holidays_lessons"] += 1
                    staff_data["busy_holidays"].add(lesson_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Таблица занятости"

    row = 1
    ws.cell(row, 1).value = (
        "Сведения о количестве занятых выходных дней с "
        f"{month_start} по {month_end} месяц {year} года"
    )
    ws.cell(row, 1).style = ExcelStyle.Header

    # Заголовки: название и ширина столбца
    headers = {
        "Имя": 40,
        "Всего занятий": 10,
        "Занятий в выходные": 15,
        "% от общего числа занятий": 15,
        "Занято выходных": 12,
        "Всего выходных": 12,
        "Среднее кол-во занятий в выходные": 15,
    }

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))

    row = 2
    column = 1

    for key, val in headers.items():
        ws.cell(row, column).value = key
        ws.cell(row, column).style = ExcelStyle.Header
        ws.cell(row, column).fill = ExcelStyle.GreyFill
        ws.column_dimensions[get_column_letter(column)].width = val
        column += 1

    row = 3

    for curr_staff in sorted(
        staff_busy_holidays.values(),
        key=lambda x: len(x.get("busy_holidays")),
        reverse=True,
    ):
        curr_name = curr_staff.get("name")
        curr_total_lessons = curr_staff.get("total_lessons")
        curr_holidays_lessons = curr_staff.get("holidays_lessons")
        curr_percent_of_lessons = curr_staff.get('holidays_lessons') / curr_staff.get('total_lessons')
        curr_busy_holidays = len(curr_staff.get("busy_holidays"))
        curr_total_holidays = len(total_holidays)
        try:
            curr_avg_lessons_holidays = (
                curr_holidays_lessons / curr_busy_holidays
            )
        except ZeroDivisionError:
            curr_avg_lessons_holidays = 0.0

        ws.cell(row, 1).value = curr_name
        ws.cell(row, 1).style = ExcelStyle.Base_No_Wrap
        ws.cell(row, 2).value = curr_total_lessons
        ws.cell(row, 2).style = ExcelStyle.Number
        ws.cell(row, 3).value = curr_holidays_lessons
        ws.cell(row, 3).style = ExcelStyle.Number
        ws.cell(row, 4).value = curr_percent_of_lessons
        ws.cell(row, 4).style = ExcelStyle.Number
        ws.cell(row, 4).number_format = '0.00%'
        ws.cell(row, 5).value = curr_busy_holidays
        ws.cell(row, 5).style = ExcelStyle.Number
        ws.cell(row, 6).value = curr_total_holidays
        ws.cell(row, 6).style = ExcelStyle.Number
        ws.cell(row, 7).value = curr_avg_lessons_holidays
        ws.cell(row, 7).style = ExcelStyle.Number
        ws.cell(row, 7).number_format = '0.00'

        row += 1

    filename = f"holidays_{year}-{month_start}-to-{month_end}-month.xlsx"

    wb.save(os.path.join(FlaskConfig.EXPORT_FILE_DIR, filename))

    return filename
