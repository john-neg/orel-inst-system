import logging
import os

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.core.classes import LoadReportProcessor
from config import FlaskConfig
from .ExcelStyles import ExcelStyle


def generate_load_report(load: LoadReportProcessor) -> str:
    """
    Формирует отчет о нагрузке кафедры в формате xlsx.

    Parameters
    ----------
        load
            экземпляр класса "LoadReportProcessor"

    Returns
    -------
        str
            название файла
    """

    wb = load_workbook(
        os.path.join(FlaskConfig.TEMPLATE_FILE_DIR, "load_report_temp.xlsx")
    )
    ws = wb.active
    ws.title = (
        f"{load.year}-{load.file_period} "
        f"{load.departments.get(load.department_id).get('short')}"
    )
    ws.cell(1, 1).value = "Кафедра " + load.departments.get(load.department_id).get(
        "full"
    )
    ws.cell(2, 1).value = f"отчет о нагрузке за {load.file_period} {load.year}"
    row = 8
    for staff_id in load.load_data:
        # Style apply
        for i in range(2, 73):
            ws.cell(row, i).style = ExcelStyle.Number
        # Prepod Name
        ws.cell(row, 1).value = load.staff_list[staff_id]
        ws.cell(row, 1).style = ExcelStyle.Base
        for l_type in load.load_data[staff_id]:
            if l_type == "lecture":
                column = 2
            elif l_type == "seminar":
                column = 8
            elif l_type == "pract":
                column = 14
            elif l_type == "group_cons":
                column = 24
                if load.load_data[staff_id][l_type]["dpo"]:
                    del load.load_data[staff_id][l_type]["dpo"]
            elif l_type == "zachet":
                column = 29
            elif l_type == "exam":
                column = 35
            elif l_type == "final_att":
                column = 59
            else:
                column = 73
            for key, val in load.load_data[staff_id][l_type].items():
                val = "" if val == 0 else val
                ws.cell(row, column).value = val
                if val and val % 1 > 0:
                    ws.cell(row, column).number_format = "0.00"
                column += 1
        ws.cell(row, 72).value = f"=SUM(B{str(row)}:BS{str(row)})"
        row += 1
    # Total
    ws.cell(row, 1).value = "Итого"
    ws.cell(row, 1).style = ExcelStyle.BaseBold
    for col in range(2, 73):
        ltr = ws.cell(row, col).column_letter
        ws.cell(row, col).value = (
            f"=IF(SUM({ltr}8:{ltr}{str(row - 1)})>0,"
            f'SUM({ltr}8:{ltr}{str(row - 1)}),"")'
        )
        ws.cell(row, col).style = ExcelStyle.Number
        ws.cell(row, col).font = Font(name="Times New Roman", size=10, bold=True)

    filename = (
        f"{load.departments.get(load.department_id).get('short')} "
        f"{load.file_period} {load.year}.xlsx"
    )
    wb.save(os.path.join(FlaskConfig.EXPORT_FILE_DIR, filename))
    logging.debug(f"Сформирован файл - отчет о нагрузке: {filename}")
    return filename
