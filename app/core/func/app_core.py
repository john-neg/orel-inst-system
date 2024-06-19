# import datetime
import json
import logging
from datetime import datetime, timedelta
from functools import lru_cache, wraps

from openpyxl import Workbook

from config import FlaskConfig


def allowed_file(filename: str) -> bool:
    """Проверяет что расширение файла в списке разрешенных."""
    return (
        "." in filename and filename.rsplit(".", 1)[1] in FlaskConfig.ALLOWED_EXTENSIONS
    )


def data_processor(table_data, dict_key: str = "id") -> dict:
    """
    Преобразует полученные данные из таблиц БД Апекс-ВУЗ.

    Parameters
    ----------
        table_data
            данные таблицы, содержащей список словарей в формате JSON
        dict_key: str
            название поля значения которого станут ключами словаря
            по умолчанию - 'id'

    Returns
    -------
        dict
            {id: {keys: values}}.
    """
    data = {}
    for d_val in table_data:
        key_val = d_val.get(dict_key)
        if key_val:
            data[int(key_val)] = d_val
    logging.debug(f"Обработаны данные. Ключ: {dict_key}")
    return data


def xlsx_iter_rows(worksheet: Workbook.active):
    """
    Чтение данных таблицы XLSX по рядам.

    Parameters
    ----------
        worksheet: Workbook.active
            активный лист таблицы
    """
    for row in worksheet.iter_rows():
        yield [cell.value for cell in row]


def xlsx_normalize(worksheet: Workbook.active, replace: dict) -> Workbook.active:
    """
    Функция для замены символов в таблице.

    Parameters
    ----------
        worksheet: Workbook.active
            активный лист таблицы
        replace: dict
            словарь для замены {'изменяемое значение': 'новое значение'}

    Returns
    -------
        worksheet: Workbook.active
            измененный лист
    """
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            for key, val in replace.items():
                cell_data = str(worksheet.cell(row, col).value).strip()
                worksheet.cell(row, col).value = cell_data.replace(key, val)
    return worksheet


def read_json_file(file_path: str) -> dict:
    """Читает файл формата json."""
    with open(file_path, encoding="utf-8") as file:
        return json.loads(file.read())


def timed_lru_cache(seconds: int, maxsize: int = 128):
    def wrapper_cache(func):
        func = lru_cache(maxsize=maxsize)(func)
        func.lifetime = timedelta(seconds=seconds)
        func.expiration = datetime.utcnow() + func.lifetime

        @wraps(func)
        def wrapped_func(*args, **kwargs):
            if datetime.utcnow() >= func.expiration:
                func.cache_clear()
                func.expiration = datetime.utcnow() + func.lifetime

            return func(*args, **kwargs)
        return wrapped_func
    return wrapper_cache
