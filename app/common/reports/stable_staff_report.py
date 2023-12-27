import logging
import os

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from app.common.reports.ExcelStyles import ExcelStyle
from config import ApeksConfig as Apeks, FlaskConfig


def generate_stable_staff_report(db_data: dict | None, busy_types: dict) -> str:
    """
    Формирует отчет о нагрузке кафедры в формате xlsx.

    Parameters
    ----------
        db_data
            данные о наличии личного состава
        busy_types
            данные о видах отвлечений

    Returns
    -------
        str
            название файла
    """

    if not db_data:
        return "no data"
    else:
        workdate = db_data.get("date")
        title = f"Строевая записка постоянного состава за {workdate}"
        wb = Workbook()
        ws = wb.active
        ws.title = workdate

        # Номера строк для объединения
        row_to_merge = []

        # Заголовок таблицы
        row = 1
        ws.cell(row, 1).value = title
        ws.cell(row, 1).style = ExcelStyle.Header
        ws.cell(row, 1).font = Font(name="Times New Roman", size=14, bold=True)
        ws.cell(row, 1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True
        )
        row_to_merge.append(row)

        # Базовые столбцы таблицы: название и ширина столбца
        row = 2
        headers = {"Подразделение": 15, "По списку": 3, "В строю": 3, "Отсутствуют": 3}

        row += 1
        addon_headers = []
        addon_values = []
        total_staff = 0
        for _, dept_type in Apeks.DEPT_TYPES.items():
            ws.cell(row, 1).value = dept_type
            ws.cell(row, 1).style = ExcelStyle.Header
            ws.cell(row, 1).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
                shrink_to_fit=True,
            )
            row_to_merge.append(row)
            row += 1
            for dept_id, dept in sorted(
                db_data["departments"].items(), key=lambda x: x[1]["name"]
            ):
                if dept.get("type") == dept_type:
                    ws.cell(row, 1).value = dept.get("name")
                    ws.cell(row, 1).style = ExcelStyle.BaseBold
                    dept_total = dept.get("total")
                    total_staff += dept_total
                    ws.cell(row, 2).value = dept_total
                    ws.cell(row, 2).style = ExcelStyle.Number
                    ws.cell(
                        row, 3
                    ).value = (f"={ws.cell(row, 2).column_letter}{row}"
                               f"-{ws.cell(row, 4).column_letter}{row}")
                    ws.cell(row, 3).style = ExcelStyle.Number
                    ws.cell(row, 4).value = (
                        sum(len(i) for i in dept["absence"].values())
                        if dept.get("absence")
                        else 0
                    )
                    ws.cell(row, 4).style = ExcelStyle.Number
                    absence_info = dept.get("absence")
                    if absence_info:
                        for header in absence_info:
                            if header not in addon_headers:
                                addon_headers.append(header)
                                addon_values.append(0)
                            header_index = addon_headers.index(header)
                            ws.cell(
                                row, len(headers) + header_index + 1
                            ).value = "\n".join(absence_info[header].values())
                            ws.cell(
                                row, len(headers) + header_index + 1
                            ).style = ExcelStyle.Number
                            addon_values[header_index] += len(
                                absence_info[header].values()
                            )
                    row += 1
        # Total
        ws.cell(row, 1).value = "Итого"
        ws.cell(row, 1).style = ExcelStyle.Header
        ws.cell(row, 1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True
        )
        row_to_merge.append(row)
        row += 1
        ws.cell(row, 1).value = "Итого"
        ws.cell(row, 1).style = ExcelStyle.BaseBold
        ws.cell(row, 2).value = total_staff
        ws.cell(row, 2).style = ExcelStyle.Number
        ws.cell(row, 3).value = total_staff - sum(addon_values)
        ws.cell(row, 3).style = ExcelStyle.Number
        ws.cell(row, 4).value = sum(addon_values)
        ws.cell(row, 4).style = ExcelStyle.Number
        column = 5
        for value in addon_values:
            ws.cell(row, column).value = value
            ws.cell(row, column).style = ExcelStyle.Number
            column += 1

        # Создаем заголовки
        row = 2
        column = 1
        for busy_type in addon_headers:
            headers[busy_types.get(busy_type)] = 13
        ws.row_dimensions[row].height = 85
        for key, val in headers.items():
            ws.cell(row, column).value = key
            ws.cell(row, column).style = ExcelStyle.Header
            ws.cell(row, column).fill = ExcelStyle.GreyFill
            text_rotation = 90 if val <= 5 else 0
            ws.cell(row, column).alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
                shrink_to_fit=True,
                textRotation=text_rotation,
            )
            ws.column_dimensions[get_column_letter(column)].width = val
            column += 1
        for row in row_to_merge:
            ws.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=len(headers)
            )

        # Формируем файл отчета
        filename = f"{title}.xlsx"
        wb.save(os.path.join(FlaskConfig.EXPORT_FILE_DIR, filename))
        logging.debug(f'Файл c расписанием "{filename}" успешно сформирован')
        return filename
