import logging
import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.common.classes.ScheduleLessonsStaff import ScheduleLessonsStaff
from config import FlaskConfig, ApeksConfig as Apeks
from .ExcelStyles import ExcelStyle


def generate_schedule_xlsx(schedule: ScheduleLessonsStaff, staff_name: str) -> str:
    """
    Формирует файл для экспорта занятий преподавателя в формате xlsx.

    Parameters
    ----------
        schedule
            экземпляр класса ScheduleLessonsStaff
        staff_name:str
            имя преподавателя

    Returns
    -------
        string
            строка с названием файла или 'no data' если список 'lessons_data' пуст
    """

    if not schedule.lessons_data:
        return "no data"
    else:
        month_name = Apeks.MONTH_DICT.get(int(schedule.month))
        filename = f"{staff_name} - {month_name} {schedule.year}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = staff_name

        row = 1
        ws.cell(row, 1).value = staff_name
        ws.cell(row, 1).style = ExcelStyle.Header
        ws.cell(row, 2).value = f"Расписание на {month_name} {schedule.year}"
        ws.cell(row, 2).style = ExcelStyle.Header

        row = 2
        column = 1

        # Заголовки: название и ширина столбца
        headers = {"Дата/время": 15, "Занятие": 65, "Место": 15, "Тема": 80}

        for key, val in headers.items():
            ws.cell(row, column).value = key
            ws.cell(row, column).style = ExcelStyle.Header
            ws.cell(row, column).fill = ExcelStyle.GreyFill
            ws.column_dimensions[get_column_letter(column)].width = val
            column += 1

        row = 3
        for l_index in range(sum(1 for _ in schedule.lessons_data)):
            ws.cell(row, 1).value = schedule.time_start(l_index).strftime(
                "%d.%m.%Y %H:%M"
            )
            ws.cell(row, 1).style = ExcelStyle.Base_No_Wrap
            ws.cell(row, 2).value = schedule.calendar_name(l_index)
            ws.cell(row, 2).style = ExcelStyle.Base_No_Wrap
            ws.cell(row, 3).value = schedule.lessons_data[l_index].get("classroom")
            ws.cell(row, 3).style = ExcelStyle.Base_No_Wrap
            ws.cell(row, 4).value = schedule.lessons_data[l_index].get("topic_name")
            ws.cell(row, 4).style = ExcelStyle.Base_No_Wrap
            row += 1

        wb.save(os.path.join(FlaskConfig.EXPORT_FILE_DIR, filename))
        logging.debug(f'Файл c расписанием "{filename}" успешно сформирован')
        return filename
