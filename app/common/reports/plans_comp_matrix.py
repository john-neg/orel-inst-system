import logging
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from app.common.classes.EducationPlan import EducationPlanCompetencies
from app.common.reports.ExcelStyles import ExcelStyle
from config import FlaskConfig, ApeksConfig as Apeks


def generate_plans_comp_matrix(plan: EducationPlanCompetencies) -> str:
    """Формирование матрицы компетенций плана в формате Excel"""

    # Сортировка из Апекс-ВУЗ (по 'left_node')
    plan_disc = {}
    for disc_id, disc in plan.plan_curriculum_disciplines.items():
        if disc.get("level") in ("1", "2", "3"):
            plan_disc[int(disc.get("left_node"))] = {
                "id": disc_id,
                "code": disc.get("code"),
                "name": disc.get("name"),
                "level": disc.get("level"),
                "type": disc.get("type"),
            }
    plan_comp = {}
    for comp in plan.plan_competencies.values():
        plan_comp[int(comp.get("left_node"))] = {
            "id": int(comp.get("id")),
            "code": comp.get("code"),
        }

    wb = Workbook()
    ws = wb.active
    ws.title = "Матрица компетенций"
    ws.cell(1, 1).value = "Код"
    ws.cell(1, 1).style = ExcelStyle.Header
    ws.column_dimensions[get_column_letter(1)].width = 15
    ws.cell(1, 2).value = "Название дисциплины"
    ws.cell(1, 2).style = ExcelStyle.Header
    ws.column_dimensions[get_column_letter(2)].width = 60

    row = 2
    for disc in sorted(plan_disc):
        ws.cell(row, 1).value = plan_disc[disc].get("code")
        ws.cell(row, 1).style = ExcelStyle.Base
        ws.cell(row, 2).value = plan_disc[disc].get("name")
        ws.cell(row, 2).style = ExcelStyle.Base
        column = 3
        for comp in sorted(plan_comp):
            ws.cell(1, column).value = plan_comp[comp].get("code")
            ws.cell(1, column).style = ExcelStyle.Header
            ws.cell(1, column).alignment = Alignment(text_rotation=90)
            ws.column_dimensions[get_column_letter(column)].width = 4
            ws.cell(row, column).style = ExcelStyle.Base
            ws.cell(row, column).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            if plan_disc[disc].get("level") != str(Apeks.DISC_LEVEL) or plan_disc[
                disc
            ].get("type") == str(Apeks.DISC_GROUP_TYPE):
                ws.cell(row, 1).style = ExcelStyle.BaseBold
                ws.cell(row, 1).fill = ExcelStyle.GreyFill
                ws.cell(row, 2).style = ExcelStyle.BaseBold
                ws.cell(row, 2).fill = ExcelStyle.GreyFill
                ws.cell(row, column).style = ExcelStyle.BaseBold
                ws.cell(row, column).fill = ExcelStyle.GreyFill
            disc_id = plan_disc[disc].get("id")
            if disc_id in plan.discipline_competencies:
                for relation in plan.discipline_competencies.get(disc_id):
                    if relation == plan_comp[comp].get("id"):
                        ws.cell(row, column).value = "+"
            column += 1
        row += 1

    filename = f"Матрица - {plan.name}.xlsx"
    wb.save(os.path.join(FlaskConfig.EXPORT_FILE_DIR, filename))
    logging.debug(f"Сформирован файл: {filename}")
    return filename
