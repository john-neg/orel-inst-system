from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime, timedelta, tzinfo

import pytz
from icalendar import Calendar, Timezone, TimezoneStandard, Event, Alarm
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.common.classes.ExcelStyle import ExcelStyle
from config import FlaskConfig, ApeksConfig as Apeks


@dataclass
class ScheduleLessonsStaff:
    """
    Класс для работы со списком занятий преподавателя за
    определенный период.

    Attributes:
    -----------
        staff_id : int | str
            id преподавателя.
        month : int | str
            месяц (число от 1 до 12).
        year : int | str
            учебный год (число 20хх).
        lessons_data : dict
            ответ GET запроса API "/api/call/schedule-schedule/staff"
            в формате JSON.
        disciplines : dict
            список дисциплин в формате {id: {'full': 'название
            дисциплины', 'short': 'сокращенное название'}}.
        load_subgroups_data: dict
            словарь для поиска 'group_id' в случае его отсутствия
            по 'subgroup_id' в формате {subgroup_id: {'group_id': val}}

    Methods:
    --------
        calendar_name(l_index: int) -> str
            выводит заголовок для занятия по его индексу в self.data.
        time_start(l_index: int) -> datetime
            выводит время начала занятия.
        time_end(l_index: int) -> datetime
            выводит время окончания занятия.
        export_ical(staff_name: str, timezone: tzinfo) -> str
            формирует файл для экспорта занятий преподавателя в формате iCal.
        export_xlsx(staff_name: str) -> str
            формирует файл для экспорта занятий преподавателя в формате xlsx.
    """

    staff_id: int | str
    month: int | str
    year: int | str
    lessons_data: list
    disciplines: dict
    load_subgroups_data: dict

    def calendar_name(self, l_index: int) -> str:
        """
        Выводит полный заголовок занятия по индексу.

        Parameters
        ----------
            l_index: int
                индекс занятия в списке 'lessons_data'
        """
        class_type_name = self.lessons_data[l_index].get("class_type_name")
        topic_code = self.lessons_data[l_index].get("topic_code")
        topic_code_fixed = f" ({topic_code})" if topic_code else ""
        discipline_id = self.lessons_data[l_index].get("discipline_id")
        try:
            short_disc_name = self.disciplines.get(int(discipline_id)).get("short")
        except AttributeError:
            logging.error(
                f"Не найдено название дисциплины по 'discipline_id': {discipline_id}"
            )
            short_disc_name = "Название дисциплины отсутствует"
        group = self.lessons_data[l_index].get("groupName")
        name = f"{class_type_name}{topic_code_fixed} {short_disc_name} {group}"
        return name

    def time_start(self, l_index: int) -> datetime:
        """
        Выводит время начала занятия

        Parameters
        ----------
            l_index: int
                индекс занятия в списке 'lessons_data'
        """
        date = datetime.strptime(
            self.lessons_data[l_index].get("date"), "%d.%m.%Y"
        ).date()
        time = datetime.strptime(
            self.lessons_data[l_index].get("lessonTime").split(" - ")[0], "%H:%M"
        ).time()
        return datetime.combine(date, time)

    def time_end(self, l_index: int) -> datetime:
        """
        Выводит время окончания занятия

        Parameters
        ----------
            l_index: int
                индекс занятия в списке 'lessons_data'
        """
        date = datetime.strptime(
            self.lessons_data[l_index].get("date"), "%d.%m.%Y"
        ).date()
        time = datetime.strptime(
            self.lessons_data[l_index].get("lessonTime").split(" - ")[1], "%H:%M"
        ).time()
        return datetime.combine(date, time)

    def export_ical(self, staff_name: str, timezone: tzinfo = Apeks.TIMEZONE) -> str:
        """
        Формирует файл для экспорта занятий преподавателя в формате iCal.

        Parameters
        ----------
            staff_name:str
                имя преподавателя
            timezone
                устанавливается timezone для календаря

        Returns
        -------
            string
                строка с названием файла или 'no data' если список 'lessons_data' пуст
        """

        if not self.lessons_data:
            return "no data"

        cal = Calendar()
        cal.add("calscale", "GREGORIAN")
        cal.add("version", "2.0")
        cal.add("prodid", "APEKS-VUZ-EXTENSION")
        cal_timezone = Timezone()
        cal_timezone.add("TZID", Apeks.TIMEZONE)
        tz_standard = TimezoneStandard()
        tz_standard.add("TZNAME", datetime.now(tz=timezone).strftime("%Z"))
        tz_standard["TZOFFSETFROM"] = datetime.now(tz=timezone).strftime("%z")
        tz_standard["TZOFFSETTO"] = datetime.now(tz=timezone).strftime("%z")
        cal_timezone.add_component(tz_standard)
        cal.add_component(cal_timezone)

        for l_index in range(len(self.lessons_data)):
            event = Event()
            event.add("dtstart", self.time_start(l_index).astimezone(pytz.utc))
            event.add("dtend", self.time_end(l_index).astimezone(pytz.utc))
            event.add("dtstamp", datetime.now().astimezone(pytz.utc))
            event.add(
                "description",
                f"т.{self.lessons_data[l_index].get('topic_code')} "
                f"{self.lessons_data[l_index].get('topic_name')}\n\n"
                f"Данные актуальны на: {datetime.now().strftime('%H:%M %d.%m.%Y')}",
            )
            event.add("location", self.lessons_data[l_index].get("classroom"))
            event.add("status", "CONFIRMED")
            event.add("summary", self.calendar_name(l_index))
            event.add(
                "uid",
                f'apeks-id-{self.lessons_data[l_index].get("id")}'
                f'journal-lesson-id-{self.lessons_data[l_index].get("journal_lesson_id")}',
            )
            group_id = self.lessons_data[l_index].get("group_id")
            subgroup_id = self.lessons_data[l_index].get("subgroup_id")
            lesson_id = self.lessons_data[l_index].get("journal_lesson_id")

            if not group_id:
                if subgroup_id:
                    group_id = self.load_subgroups_data.get(subgroup_id).get("group_id")

            if group_id and lesson_id:
                event.add(
                    "url",
                    f"{Apeks.URL}/student/journal/view?group_id={group_id}"
                    f"&lesson_id={lesson_id}",
                )
            alarm = Alarm()
            alarm.add("action", "DISPLAY")
            alarm.add("description", "Напоминание")
            alarm.add("trigger", timedelta(minutes=-30))
            event.add_component(alarm)
            cal.add_component(event)

        month_name = Apeks.MONTH_DICT.get(int(self.month))
        filename = f"{staff_name} - {month_name} {self.year}.ics"

        with open(
            f"{FlaskConfig.EXPORT_FILE_DIR}{filename}",
            "wb",
        ) as f:
            f.write(cal.to_ical())
            f.close()
            logging.debug(f'Файл "{filename}" успешно сформирован')
        return filename

    def export_xlsx(self, staff_name: str) -> str:
        """
        Формирует файл для экспорта занятий преподавателя в формате xlsx.

        Parameters
        ----------
            staff_name:str
                имя преподавателя

        Returns
        -------
            string
                строка с названием файла или 'no data' если список 'lessons_data' пуст
        """

        if not self.lessons_data:
            return "no data"
        else:
            month_name = Apeks.MONTH_DICT.get(int(self.month))
            filename = (
                f"{staff_name} - {month_name} {self.year}.xlsx"
            )
            wb = Workbook()
            ws = wb.active
            ws.title = staff_name

            row = 1
            ws.cell(row, 1).value = staff_name
            ws.cell(row, 1).style = ExcelStyle.Header
            ws.cell(
                row, 2
            ).value = f"Расписание на {month_name} {self.year}"
            ws.cell(row, 2).style = ExcelStyle.Header

            row = 2
            column = 1

            # Заголовки: название и ширина столбца
            headers = {"Дата/время": 15, "Занятие": 65, "Место": 15, "Тема": 80}

            for key, val in headers.items():
                ws.cell(row, column).value = key
                ws.cell(row, column).style = ExcelStyle.Header
                ws.cell(row, column).fill = ExcelStyle.GreyFill
                ws.column_dimensions[get_column_letter(column)].width = val
                column += 1

            row = 3
            for l_index in range(len(self.lessons_data)):
                ws.cell(row, 1).value = self.time_start(l_index).strftime(
                    "%d.%m.%Y %H:%M"
                )
                ws.cell(row, 1).style = ExcelStyle.Base_No_Wrap
                ws.cell(row, 2).value = self.calendar_name(l_index)
                ws.cell(row, 2).style = ExcelStyle.Base_No_Wrap
                ws.cell(row, 3).value = self.lessons_data[l_index].get("classroom")
                ws.cell(row, 3).style = ExcelStyle.Base_No_Wrap
                ws.cell(row, 4).value = self.lessons_data[l_index].get("topic_name")
                ws.cell(row, 4).style = ExcelStyle.Base_No_Wrap
                row += 1

            wb.save(FlaskConfig.EXPORT_FILE_DIR + filename)
            logging.debug(f'Файл "{filename}" успешно сформирован')
            return filename
