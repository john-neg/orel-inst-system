import re
from dataclasses import dataclass

from openpyxl import load_workbook

from common.classes.EducationPlan import (
    EducationPlanCompetencies,
    EducationPlanIndicators,
)
from common.func.app_core import xlsx_iter_rows, xlsx_normalize
from config import ApeksConfig as Apeks
from plans.func import get_competency_code


@dataclass
class MatrixFileProcessor:
    """
    Класс для обработки файла матрицы компетенций.

    :parameter:

    """

    file: str

    def __post_init__(self):
        self.file_list = self.xlsx_file_to_list()
        self.file_dict = self.xlsx_file_to_dict()
        self.matrix_file_comp = self.matrix_file_comp()

    def xlsx_file_to_list(self) -> list:
        """
        Преобразует xlsx файл в список.

        :return: список [[row], [row], ...]
        """
        wb = load_workbook(self.file)
        ws = xlsx_normalize(wb.active, Apeks.COMP_REPLACE_DICT)
        file_list = list(xlsx_iter_rows(ws))
        return file_list

    def xlsx_file_to_dict(self) -> dict:
        """
        Обрабатывает данные из xlsx файла, преобразуя их в словарь.

        :return: {discipline_name: [comp_match_list]}
        """
        file_dict = {}
        matrix = self.file_list
        for row in range(Apeks.MATRIX_COMP_ROW + 1, len(matrix)):
            disc_name = matrix[row][Apeks.MATRIX_DISC_COL]
            file_dict[disc_name] = []
            for col in range(Apeks.MATRIX_DISC_COL + 1, len(matrix[row])):
                if matrix[row][col] == "+":
                    disc_comp_data = self.file_list[Apeks.MATRIX_COMP_ROW][col]
                    file_dict[disc_name].append(disc_comp_data)
        return file_dict

    def matrix_file_comp(self) -> set:
        """
        Выводит коды всех компетенций файла матрицы.

        :return: set {'comp_codes'}
        """

        file_comp = set()
        for col in range(
            Apeks.MATRIX_DISC_COL + 1, len(self.file_list[Apeks.MATRIX_COMP_ROW])
        ):
            file_comp.add(
                get_competency_code(self.file_list[Apeks.MATRIX_COMP_ROW][col])
            )
        return file_comp

    def get_file_report_data(self):
        """

        :return:
        """
        report_data = {}
        indicator_errors = set()
        ind_regex = re.compile(Apeks.COMP_FROM_IND_REGEX)
        for disc in self.file_dict:
            report_data[disc] = {
                "comp_list": [],
                "knowledge": [],
                "abilities": [],
                "ownerships": [],
            }
            for ind in self.file_dict[disc]:
                comp_code = get_competency_code(ind)
                if comp_code not in report_data[disc]["comp_list"]:
                    report_data[disc]["comp_list"].append(comp_code)
                try:
                    ind_code, ind_val = re.split(
                        Apeks.FULL_CODE_SPLIT_REGEX, ind, 1
                    )
                    ind_separator = ind_regex.search(ind_code).group()
                except AttributeError:
                    indicator_errors.add(ind)
                else:
                    if ind_separator and ind_separator in Apeks.INDICATOR_CODES:
                        ind_type = Apeks.INDICATOR_CODES.get(ind_separator)
                        report_data[disc].get(ind_type).append(f"- {ind_val} ({ind_code})")
                    else:
                        indicator_errors.add(ind)
        return report_data, indicator_errors


@dataclass
class MatrixSimpleProcessor(MatrixFileProcessor):
    """
    Класс для обработки файла с простой матрицей компетенций и подготовки данных
    для загрузки в учебные планы и рабочие программы плана.

    Attributes:
    ----------
        plan:
            экземпляр класса EducationPlanCompetencies
        file: str
            полный путь к файлу со списком компетенций

    Methods:
    -------


    """

    plan: EducationPlanCompetencies

    def __post_init__(self) -> None:
        super().__post_init__()
        self.plan_competencies = {
            comp.get("code"): key for key, comp in self.plan.plan_competencies.items()
        }
        self.disciplines = self.plan.plan_curriculum_disciplines
        self.indicator_errors = set()
        self.matrix_match_data = self.get_matrix_match_data()

    def comp_not_in_file(self) -> set:
        """
        Выводит коды компетенций плана, отсутствующие в файле.

        :return: set {'comp_codes'}
        """
        plan_comp = set(self.plan_competencies)
        file_comp = self.matrix_file_comp
        return plan_comp.difference(file_comp)

    def comp_not_in_plan(self) -> set:
        """
        Выводит коды компетенций файла, отсутствующие в плане.

        :return: set {'comp_codes'}
        """
        plan_comp: set = set(self.plan_competencies)
        file_comp = self.matrix_file_comp
        return file_comp.difference(plan_comp)

    def get_match_data_template(self) -> dict:
        """
        Возвращает шаблон словаря заполненный данными из плана, готовый для
        для заполнения связанными данными файла.

        Returns
        -------
            dict
                {"discipline_name": {"id": disc_id, "code": disc_code,
                "left_node": left_node}.
        """

        disciplines = sorted(
            self.disciplines.values(), key=lambda d: int(d["left_node"])
        )
        match_data = {
            val.get("name"): {
                "id": val.get("id"),
                "code": val.get("code"),
                "left_node": val.get("left_node"),
            }
            for val in disciplines
            if val.get("level") == str(Apeks.DISC_LEVEL)
            and val.get("type") != str(Apeks.DISC_TYPE)
        }
        return match_data

    def get_matrix_match_data(self) -> dict:
        """
        Сопоставляет данные из плана и файла для загрузки связей
        дисциплин и компетенций из простой матрицы.

        Returns
        -------
            dict
                {"discipline_name": {"id": disc_id, "code": disc_code,
                "left_node": left_node, "comps": {"comp_name": comp_id}}.
        """
        matrix_match_data = self.get_match_data_template()
        for disc in matrix_match_data:
            if disc in self.file_dict:
                matrix_match_data[disc]["comps"] = {}
                for comp_code in self.file_dict[disc]:
                    if comp_code in self.plan_competencies:
                        matrix_match_data[disc]["comps"][
                            comp_code
                        ] = self.plan_competencies[comp_code]
        return matrix_match_data


class MatrixIndicatorProcessor(MatrixSimpleProcessor):
    """
    Класс для обработки файла с простой матрицей компетенций и подготовки данных
    для загрузки в учебные планы и рабочие программы плана.

    Attributes:
    ----------
        plan:
            экземпляр класса EducationPlanCompetencies
        file: str
            полный путь к файлу со списком компетенций

    """

    plan: EducationPlanIndicators

    def __post_init__(self) -> None:
        super().__post_init__()

    def get_matrix_match_data(self) -> dict:
        """
        Сопоставляет данные из плана и файла для загрузки связей
        дисциплин и компетенций из матрицы с индикаторами.

        :return: dict - {"discipline_name": {"id": disc_id, "code": disc_code,
            "left_node": left_node, "comps": {"comp_name": {"id": comp_id,
            "knowledge": [val], "abilities": [val], "ownerships":[val]}}}}.
        """
        matrix_match_data = self.get_match_data_template()
        ind_regex = re.compile(Apeks.COMP_FROM_IND_REGEX)
        for disc in matrix_match_data:
            disc_name = f"{matrix_match_data[disc].get('code')} {disc}"
            if disc_name in self.file_dict:
                matrix_match_data[disc]["comps"] = {}
                for ind in self.file_dict[disc_name]:
                    comp_code = get_competency_code(ind)
                    try:
                        ind_code, ind_val = re.split(
                            Apeks.FULL_CODE_SPLIT_REGEX, ind, 1
                        )
                        ind_separator = ind_regex.search(ind_code).group()
                    except AttributeError:
                        self.indicator_errors.add(ind)
                    else:
                        if ind_separator and comp_code in self.plan_competencies:
                            ind_type = Apeks.INDICATOR_CODES.get(ind_separator)
                            comp_data = matrix_match_data[disc]["comps"].setdefault(
                                comp_code,
                                {
                                    "id": self.plan_competencies[comp_code],
                                    "knowledge": [],
                                    "abilities": [],
                                    "ownerships": [],
                                },
                            )
                            comp_data[ind_type].append(f"- {ind_val} ({ind_code})")
                        else:
                            self.indicator_errors.add(ind)
        return matrix_match_data

    def program_load_data(self):
        """
        Данные для добавления, обновления и удаления информации в рабочих
        программах.

        :return: program_level_add - [список словарей для добавления уровней
            компетенций в рабочие программы], program_level_edit -
            {work_program_id: {данные для обновления существующих уровней
            компетенций }}, program_competency_add - [список словарей для
            загрузки сведений о компетенциях в рабочие программы]
        """
        program_level_add = []
        program_level_edit = {}
        program_competency_add = []
        comp_parameters = Apeks.MM_WORK_PROGRAMS_COMPETENCIES_DATA
        for wp, val in self.plan.work_programs_data.items():
            # Создаем параметры (знать, уметь, владеть...) для загрузки
            # раздел в уровни сформированности рабочей программы
            competency_param_data = {param: [] for param in comp_parameters}
            wp_name = val.get("name")
            match_data = self.matrix_match_data.get(wp_name)
            # Формируем сведения о компетенциях рабочей программы
            if match_data and match_data.get("comps"):
                for comp_name, comp_data in match_data.get("comps").items():
                    for p_name, p_id in comp_parameters.items():
                        if comp_data.get(p_name):
                            comp_add = {
                                "work_program_id": wp,
                                "competency_id": comp_data.get("id"),
                                "field_id": p_id,
                                "value": ";\n".join(comp_data.get(p_name)),
                            }
                            program_competency_add.append(comp_add)
                            competency_param_data[p_name].extend(comp_data.get(p_name))

            program_level = val["competency_levels"].get(Apeks.BASE_COMP_LEVEL)
            # Формируем сведения для обновления уровня сформированности компетенций
            if program_level:
                level_edit = {
                    "knowledge": ";\n".join(competency_param_data.get("knowledge")),
                    "abilities": ";\n".join(competency_param_data.get("abilities")),
                    "ownerships": ";\n".join(competency_param_data.get("ownerships")),
                }
                if program_level.get("semester_id") != val["control_works"].get(
                    "semester_id"
                ):
                    level_edit["semester_id"] = val["control_works"].get("semester_id")
                if program_level.get("control_type_id") != val["control_works"].get(
                    "control_type_id"
                ):
                    level_edit["control_type_id"] = val["control_works"].get(
                        "control_type_id"
                    )
                program_level_edit[wp] = level_edit
            else:
                # Формируем сведения для создания (если нет) уровня
                # сформированности компетенций
                if val["control_works"]:
                    level_add = {
                        "work_program_id": wp,
                        "level": Apeks.BASE_COMP_LEVEL,
                        "semester_id": val["control_works"].get("semester_id"),
                        "control_type_id": val["control_works"].get("control_type_id"),
                        "knowledge": ";\n".join(competency_param_data.get("knowledge")),
                        "abilities": ";\n".join(competency_param_data.get("abilities")),
                        "ownerships": ";\n".join(
                            competency_param_data.get("ownerships")
                        ),
                        "level1": "",
                        "level2": "",
                        "level3": "",
                    }
                    program_level_add.append(level_add)
        return program_level_add, program_level_edit, program_competency_add
