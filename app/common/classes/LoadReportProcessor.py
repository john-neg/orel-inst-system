from __future__ import annotations

import logging
from dataclasses import dataclass

from openpyxl import load_workbook
from openpyxl.styles import Font

from app.common.classes.ExcelStyle import ExcelStyle
from app.common.classes.LessonsData import LessonsData
from config import FlaskConfig, ApeksConfig as Apeks


@dataclass
class LoadReportProcessor(LessonsData):
    """
    Класс для формирования отчета об учебной нагрузке по кафедрам

    Attributes:
    ----------
        year: int | str
            учебный год (число 20xx).
        month_start: int | str
            начальный месяц (число 1-12).
        month_end: int | str
            конечный месяц (число 1-12).
        department_id: int
            id кафедры
        departments: dict
            список кафедр, словарь в формате {id: {'full':
            'название кафедры', 'short': 'сокращенное название'}}
        department_staff: dict
            словарь преподавателей, работавших в подразделении
            ('department_id') в указанный период. {id: 'short_name'}

    Methods:
    -------
        add_load (staff_id: int, l_type: str, s_type: str, value: float) -> None:
            добавляет нагрузку в хранилище 'load_data'
        get_load(staff_id: int) -> dict:
            возвращает нагрузку преподавателя их хранилища 'load_data'
        process_load_data() -> None:
            рассчитывает нагрузку по преподавателям
        generate_report() -> str:
            формирует файл отчета о нагрузке в Excel, возвращает имя файла.
    """

    year: int | str
    month_start: int | str
    month_end: int | str
    department_id: int
    departments: dict
    department_staff: dict

    def __post_init__(self) -> None:
        super().__post_init__()
        self.staff_list = self.department_staff
        self.load_data = {}
        self.unprocessed = []
        self.process_load_data()
        self.file_period = (
            Apeks.MONTH_DICT[int(self.month_start)]
            if self.month_start == self.month_end
            else f"{Apeks.MONTH_DICT[int(self.month_start)]}-"
            f"{Apeks.MONTH_DICT[int(self.month_end)]}"
        )

    def add_load(
        self, staff_id: int, l_type: str, s_type: str, value: float = 2
    ) -> None:
        """
        Добавляет нагрузку в хранилище 'load_data'

        Parameters
        ----------
            staff_id: int
                id преподавателя
            l_type: str
                тип занятия
            s_type: str
                тип обучающегося
            value: float = 2
                значение
        """
        self.load_data[staff_id][l_type][s_type] += value

    def get_load(self, staff_id: int) -> dict:
        """
        Возвращает нагрузку преподавателя их хранилища 'load_data'

        Parameters
        ----------
            staff_id: int
                id преподавателя

        Returns
        -------
            dict
                {'short_name': {'exam': {'adj': 0,
                               'dpo': 0,
                               'och': 0,
                               'prof_pod': 0,
                               'zo_high': 0,
                               'zo_mid': 0},
        """
        staff_load = {self.staff_list[staff_id]: self.load_data[staff_id]}
        return staff_load

    def process_load_data(self) -> None:
        """Рассчитывает нагрузку по преподавателям"""
        lesson_types = Apeks.LOAD_LESSON_TYPES + Apeks.LOAD_CONTROL_TYPES
        for staff_id in self.staff_list:
            self.load_data[staff_id] = {}
            for l_type in lesson_types:
                self.load_data[staff_id][l_type] = {}
                for s_type in Apeks.LOAD_STUDENT_TYPES:
                    self.load_data[staff_id][l_type][s_type] = 0
        for lesson in self.structured_lessons:
            department_id = lesson.get("department_id")
            if department_id:
                if department_id == int(self.department_id):
                    staff_id = lesson.get("staff_id")
                    l_type = self.get_lesson_type(lesson)
                    s_type = self.get_student_type(lesson)
                    if staff_id and l_type and s_type:
                        if l_type in Apeks.LOAD_LESSON_TYPES:
                            self.add_load(staff_id, l_type, s_type)
            else:
                logging.warning(f"Необработанное занятие - {lesson}")
                self.unprocessed.append(lesson)
        for control in self.control_lessons:
            department_id = control.get("department_id")
            if department_id:
                if department_id == int(self.department_id):
                    staff_id = control.get("staff_id")
                    l_type = self.get_lesson_type(control)
                    s_type = self.get_student_type(control)
                    if staff_id and l_type and s_type:
                        if l_type in Apeks.LOAD_CONTROL_TYPES:
                            value = self.get_control_hours(control)
                            self.add_load(staff_id, l_type, s_type, value)
            else:
                logging.warning(f"Необработанное занятие типа контроль - {control}")
                self.unprocessed.append(control)

    def generate_report(self) -> str:
        """
        Формирует отчет о нагрузке в Excel.

        Returns
        -------
            str
                название файла
        """

        wb = load_workbook(FlaskConfig.TEMP_FILE_DIR + "load_report_temp.xlsx")
        ws = wb.active
        ws.title = (
            f"{self.year}-{self.file_period} "
            f"{self.departments.get(self.department_id).get('short')}"
        )
        ws.cell(1, 1).value = "Кафедра " + self.departments.get(self.department_id).get(
            "full"
        )
        ws.cell(2, 1).value = f"отчет о нагрузке за {self.file_period} {self.year}"
        row = 8
        for staff_id in self.load_data:
            # Style apply
            for i in range(2, 73):
                ws.cell(row, i).style = ExcelStyle.Number
            # Prepod Name
            ws.cell(row, 1).value = self.staff_list[staff_id]
            ws.cell(row, 1).style = ExcelStyle.Base
            for l_type in self.load_data[staff_id]:
                if l_type == "lecture":
                    column = 2
                elif l_type == "seminar":
                    column = 8
                elif l_type == "pract":
                    column = 14
                elif l_type == "group_cons":
                    column = 24
                    if self.load_data[staff_id][l_type]["dpo"]:
                        del self.load_data[staff_id][l_type]["dpo"]
                elif l_type == "zachet":
                    column = 29
                elif l_type == "exam":
                    column = 35
                elif l_type == "final_att":
                    column = 59
                else:
                    column = 73
                for key, val in self.load_data[staff_id][l_type].items():
                    val = "" if val == 0 else val
                    ws.cell(row, column).value = val
                    if val and val % 1 > 0:
                        ws.cell(row, column).number_format = "0.00"
                    column += 1
            ws.cell(row, 72).value = f"=SUM(B{str(row)}:BS{str(row)})"
            row += 1
        # Total
        ws.cell(row, 1).value = "Итого"
        ws.cell(row, 1).style = ExcelStyle.BaseBold
        for col in range(2, 73):
            ltr = ws.cell(row, col).column_letter
            ws.cell(row, col).value = (
                f"=IF(SUM({ltr}8:{ltr}{str(row - 1)})>0,"
                f'SUM({ltr}8:{ltr}{str(row - 1)}),"")'
            )
            ws.cell(row, col).style = ExcelStyle.Number
            ws.cell(row, col).font = Font(name="Times New Roman", size=10, bold=True)

        filename = (
            f"{self.departments.get(self.department_id).get('short')} "
            f"{self.file_period} {self.year}.xlsx"
        )
        wb.save(FlaskConfig.EXPORT_FILE_DIR + filename)
        logging.debug(f"Сформирован файл - отчет о нагрузке: {filename}")
        return filename
