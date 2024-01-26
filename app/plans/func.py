import re

from openpyxl import load_workbook

from app.core.classes.EducationPlan import (
    EducationPlanCompetencies,
    EducationPlanIndicators,
)
from app.core.func.api_get import (
    check_api_db_response,
    api_get_db_table,
)
from app.core.func.app_core import xlsx_iter_rows, xlsx_normalize, data_processor
from app.core.func.education_plan import get_plan_curriculum_disciplines, \
    get_plan_discipline_competencies, plan_disciplines_competencies_del, \
    plan_competencies_del
from app.core.func.work_program import get_work_programs_data, \
    work_programs_competencies_del
from config import ApeksConfig as Apeks


def comps_file_processing(file: str) -> list:
    """
    Обработка загруженного файла c компетенциями.

    Parameters
    ----------
        file: str
            полный путь к файлу со списком компетенций

    Returns
    -------
        list
            нормализованный список компетенций из файла без первой строки
    """

    wb = load_workbook(file)
    ws = wb.active
    ws = xlsx_normalize(ws, Apeks.COMP_REPLACE_DICT)
    comps = list(xlsx_iter_rows(ws))
    del comps[0]
    return comps


def get_competency_code(indicator) -> str:
    """
    Выводит код компетенций индикатора.

    :return: код компетенции
    """
    comp_code = re.split(Apeks.COMP_FROM_IND_REGEX, indicator, 1)[0]
    if len(comp_code) > 12:
        comp_code = re.split(Apeks.FULL_CODE_SPLIT_REGEX, indicator, 1)[0]
    return comp_code


async def get_plan_competency_instance(plan_id: int | str) -> EducationPlanCompetencies:
    """
    Возвращает экземпляр класса 'EducationPlanCompetencies' с
    данными, необходимыми для работы приложения 'plans'
    """
    plan_disciplines = await get_plan_curriculum_disciplines(plan_id, disc_filter=False)
    plan = EducationPlanCompetencies(
        education_plan_id=plan_id,
        plan_education_plans=await check_api_db_response(
            await api_get_db_table(Apeks.TABLES.get("plan_education_plans"), id=plan_id)
        ),
        plan_curriculum_disciplines=plan_disciplines,
        plan_competencies=data_processor(
            await check_api_db_response(
                await api_get_db_table(
                    Apeks.TABLES.get("plan_competencies"), education_plan_id=plan_id
                )
            )
        ),
        discipline_competencies=await get_plan_discipline_competencies(
            [*plan_disciplines]
        ),
    )
    return plan


async def get_plan_indicator_instance(plan_id: int | str) -> EducationPlanIndicators:
    """
    Возвращает экземпляр класса 'EducationPlanIndicators' с
    данными, необходимыми для работы модуля Матрица с индикаторами
    приложения 'plans'.
    """
    plan_disciplines = await get_plan_curriculum_disciplines(plan_id)
    work_programs_data = await get_work_programs_data(
        curriculum_discipline_id=[*plan_disciplines], competencies=True
    )
    plan = EducationPlanIndicators(
        education_plan_id=plan_id,
        plan_education_plans=await check_api_db_response(
            await api_get_db_table(Apeks.TABLES.get("plan_education_plans"), id=plan_id)
        ),
        plan_curriculum_disciplines=plan_disciplines,
        plan_competencies=data_processor(
            await check_api_db_response(
                await api_get_db_table(
                    Apeks.TABLES.get("plan_competencies"), education_plan_id=plan_id
                )
            )
        ),
        discipline_competencies=await get_plan_discipline_competencies(
            [*plan_disciplines]
        ),
        work_programs_data=work_programs_data,
    )
    return plan


async def plan_competencies_data_cleanup(
    plan_id: int | str,
    plan_disciplines: list | tuple | dict,
    plan_comp: bool = True,
    relations: bool = True,
    work_program: bool = True,
) -> str:
    """
    Удаляет данные о компетенциях и связях дисциплин и компетенций из
    учебного плана и рабочих программ.

    Parameters
    ----------
        plan_id: int | str
            id учебного плана
        plan_disciplines: list | tuple | dict
            id дисциплин учебного плана
        plan_comp: bool
            удалять компетенции
        relations: bool
            удалять связи
        work_program: bool
            удалять из рабочих программ

    Returns
    -------
        str
            сведения о количестве удаленных элементов
    """
    message = ["Произведена очистка компетенций. ", "Количество удаленных записей: "]
    if work_program:
        work_programs_data = await get_work_programs_data(
            curriculum_discipline_id=[*plan_disciplines]
        )
        wp_resp = await work_programs_competencies_del(
            work_program_id=[*work_programs_data]
        )
        message.append(f"в рабочих программах - {wp_resp.get('data')},")
    if relations:
        disc_resp = await plan_disciplines_competencies_del(
            curriculum_discipline_id=[*plan_disciplines]
        )
        message.append(f"связей с дисциплинами - {disc_resp.get('data')},")
    if plan_comp:
        plan_resp = await plan_competencies_del(education_plan_id=plan_id)
        message.append(f"компетенций плана - {plan_resp.get('data')}.")
    return " ".join(message)
