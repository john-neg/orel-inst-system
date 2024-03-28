import os
import asyncio
import datetime as dt
from openpyxl import load_workbook

from app.core.services.apeks_db_state_departments_service import get_db_apeks_state_departments_service
from app.core.services.apeks_db_state_staff_field_data_service import get_apeks_db_state_staff_field_data_service
from app.core.services.apeks_db_state_staff_history_service import get_db_apeks_state_staff_history_service
from app.core.services.apeks_db_state_staff_service import get_apeks_db_state_staff_service, process_state_staff_data
from config import BASEDIR

"""
для заполнения базы апекс-вуза телефонными номерами
"""

# для работы с xlsx файлом
FILE_DIR = os.path.join(BASEDIR, 'tools', 'phonebook')
wb = load_workbook(filename=os.path.join(FILE_DIR, 'phones.xlsx'))
wb.active = 0
ws = wb.active

# список сотрудников из Апекс-ВУЗ
# дата на которую получаем список сотрудников
working_date = dt.date.today()
# получаем список подразделений
departments_service = get_db_apeks_state_departments_service()
departments = asyncio.run(departments_service.get_departments())
# служба для получения сотрудников на определенную дату
staff_history_service = get_db_apeks_state_staff_history_service()
# служба для получения полей профиля сотрудников
staff_fields_data_service = get_apeks_db_state_staff_field_data_service()
# staff_fields_data = data_processor(asyncio.run(staff_fields_data_service.list()))

# пробегаемся по всем подразделениям
for dep_id in departments.keys():
    # список сотрудников в подразделении
    staff_history = asyncio.run(staff_history_service.get_staff_for_date(working_date, department_id=dep_id))
    staff_ids = {
        item.get("staff_id"): item
        for item in staff_history
        if item.get("vacancy_id") and item.get("value") == "1"
    }
    state_staff_service = get_apeks_db_state_staff_service()
    state_staff = process_state_staff_data(
        asyncio.run(state_staff_service.get(id=staff_ids.keys()))
    )
    for key, val in state_staff.items():
        print(key, ' ', val['full'], end=' - ')
        for row in ws.iter_rows():
            # если данные и первой и во второй ячейке строки таблицы, то это запись о пользователе
            if row[0].value and row[1].value:
                if row[1].value == val['full']:
                    field_data = asyncio.run(staff_fields_data_service.get(staff_id=key, field_id=1))
                    numbers = ''
                    if row[3].value:
                        numbers = numbers + str(row[3].value)
                    if row[3].value and row[4].value:
                        numbers = numbers + ', '
                    if row[4].value:
                        numbers = numbers + str(row[4].value)
                    if len(field_data):  # если запись уже есть, то обновляем
                        asyncio.run(staff_fields_data_service.update({
                            'staff_id': key,
                            'field_id': '1'
                        }, {
                            'data': numbers
                        }))
                        print('updated', end=' - ')
                    else:  # если нет, создаем новую
                        asyncio.run(staff_fields_data_service.create(
                            staff_id=key, field_id=1, data=numbers
                        ))
                        print('created', end=' - ')
                    print(numbers)
                    break
        else:
            print('not in xlsx file!')
            pass

wb.close()
